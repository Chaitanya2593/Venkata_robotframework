'''

Created on March 2018

@author: Venkata

'''

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from AllFolders_location import cmtconfigpropFolder
from selenium import webdriver
from os import listdir
from os.path import isfile, join
from Crypto.Cipher import AES
from Crypto import Random
import base64
import random
import datetime
import sys
import string
import unicodecsv
import ConfigParser
import json
import time
import glob
import shutil
import re
import os
import logging
import hashlib

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BLOCK_SIZE = 16
pad = lambda s: s + (BLOCK_SIZE - len(s) % BLOCK_SIZE) * chr(BLOCK_SIZE - len(s) % BLOCK_SIZE)
unpad = lambda s: s[:-ord(s[len(s) - 1:])]

class CommonUtilsClass(object):

    def __init__(self):
        '''
        this is a constructor of class
        '''

    def get_val_from_config_propfile(self, configprop_filename, section_name, key_name):
        config_p = ConfigParser.RawConfigParser()
        prop_file_path = cmtconfigpropFolder + configprop_filename
        config_p.read(prop_file_path)
        val_from_propfile = config_p.get(section_name, key_name)
        return val_from_propfile

    def set_val_for_key_in_config_propfile(self, configprop_filename, section_name, key_name, key_value):
        config_p = ConfigParser.SafeConfigParser()
        prop_file_path = cmtconfigpropFolder + configprop_filename
        config_p.read(prop_file_path)
        config_p.set(section_name, key_name, str(key_value))
        with open(prop_file_path, 'w') as file_P:
            config_p.write(file_P)

    def read_json_data_for_config_prop_dict(self, json_object, primary_key_name, inner_ky):
        for each_obj in json_object:
            if isinstance(each_obj, dict):
                for ky1, vl1 in each_obj.items():
                    if (ky1 == primary_key_name):
                        if isinstance(vl1, list):
                            for lst_items in vl1:
                                if isinstance(lst_items, dict):
                                    for ky2, vl2 in lst_items.items():
                                        if (ky2 == inner_ky):
                                            val_to_b_returned = vl2
                                            return val_to_b_returned


    def read_jira_jsonfile(self, json_code):
        jiraoutput_dict = {}
        json_code = json.loads(json_code, encoding='utf-8')
        for each_obj in json_code:
            if isinstance(each_obj, dict):
                for item in each_obj.items():
                    jiraoutput_dict[each_obj['key']] = each_obj['id']
        return jiraoutput_dict

    # Common
    def str_time_prop(self, start, end, format, prop):
        stime = time.mktime(time.strptime(start, format))
        etime = time.mktime(time.strptime(end, format))
        ptime = stime + prop * (etime - stime)
        return time.strftime(format, time.localtime(ptime))

    def change_case(self, string_to_change, format_type_number):
        '''
        #'added on 19 feb 2018' by xsmaddurve
        string_to_change= "aAAAAA"
        :param string_to_change:
        :param format_type_number:
        :return:
        '''
        if format_type_number == 1:
            return string_to_change[0].lower() + string_to_change[1:len(string_to_change)].upper()

    def date_format_convert(self, input_date, format_type_number,
                            delta_varient='0'):  # input format should be dd/mm/yyyy HH:MM
        '''
        :param input_date:
        :param format_type_number:
        :param delta_varient:
        :return: Returns date in
                    1. "%B %d, %Y :%M:"
                    2. %Y/%m/%d %I:%M:%S %p"
        '''

        if format_type_number == 1:  # output as "%B %d, %Y %H:%M"
            strptime_input_date = time.strptime(str(input_date), "%d/%m/%Y %H:%M")
            date_strp_format = datetime.datetime(strptime_input_date.tm_year, strptime_input_date.tm_mon,
                                                 strptime_input_date.tm_mday, strptime_input_date.tm_hour,
                                                 strptime_input_date.tm_min)
            output_datetime = (date_strp_format - datetime.timedelta(
            minutes=float(delta_varient)))  # by default delta_varient set to zero
            final_datetime = time.strptime(str(output_datetime), "%Y-%m-%d %H:%M:%S")
            return time.strftime("%B %d, %Y :%M:", final_datetime)
        elif format_type_number == 2:  # output as "%Y/%m/%d %I:%M:%S %p"
            strptime_input_date = time.strptime(str(input_date), "%d/%m/%Y %H:%M")
            date_strp_format = datetime.datetime(strptime_input_date.tm_year, strptime_input_date.tm_mon,
                                                 strptime_input_date.tm_mday, strptime_input_date.tm_hour,
                                                 strptime_input_date.tm_min)
            output_datetime = (date_strp_format - datetime.timedelta(
            minutes=float(delta_varient)))  # by default delta_varient set to zero
            final_datetime = time.strptime(str(output_datetime), "%Y-%m-%d %H:%M:%S")
            return time.strftime("%Y/%m/%d %I:%M:%S %p", final_datetime)
        elif format_type_number == 'Current_time':
            now = input_date
            return now.strftime("%Y/%m/%d %I:%M:%S %p")
        elif format_type_number == 3: # Extracting the date from the SFMC timestamp
            try:
                strptime_input_date = time.strptime(str(input_date), "%m/%d/%Y %H:%M:%S %p")
                date_strp_format = datetime.datetime(strptime_input_date.tm_year, strptime_input_date.tm_mon,
                                                     strptime_input_date.tm_mday, strptime_input_date.tm_hour,
                                                     strptime_input_date.tm_min)
                final_datetime = time.strptime(str(date_strp_format), "%Y-%m-%d %H:%M:%S")
                return time.strftime("%Y/%m/%d", final_datetime)
            except ValueError:
                return ""
        elif format_type_number == 4: # Extracting the date from the code timestamp
            try:
                strptime_input_date = time.strptime(str(input_date), "%Y/%m/%d %I:%M:%S %p")
                date_strp_format = datetime.datetime(strptime_input_date.tm_year, strptime_input_date.tm_mon,
                                                     strptime_input_date.tm_mday, strptime_input_date.tm_hour,
                                                     strptime_input_date.tm_min)
                final_datetime = time.strptime(str(date_strp_format), "%Y-%m-%d %H:%M:%S")
                return time.strftime("%Y/%m/%d", final_datetime)
            except ValueError:
                return ""
    # Common
    def random_date(self, start, end, prop):
        return self.str_time_prop(start, end, '%Y/%m/%d %I:%M:%S %p', prop)

    def remove_list_items(self, values_of_the_list, items_to_delete):
        return [value for value in values_of_the_list if value not in items_to_delete]

    # Returns a random alphanumeric string of length 'length'
    # Common
    def random_sequences(self, random_sequence_type, length_of_value=None, no_of_fake_values=None,
                         sequence_required=None, string_to_be_concat=None):
        '''
        :param random_sequence_type:
        :param length_of_value:
        :param no_of_fake_values:
        :param sequence_required:
        :param string_to_be_concat:
        :return:
        '''
        values_of_the_list = [] 
        counter = 1
        while (counter-1 < int(no_of_fake_values)):
            if counter > 1000:
                break
            key = ''
            if (random_sequence_type.upper() == 'ALPHANUMERIC'):
                for i in range(length_of_value):
                    key += random.choice(string.ascii_lowercase + string.digits)

            elif (random_sequence_type.upper() == 'NUMERIC'):
                for i in range(length_of_value):
                    key += random.choice(string.digits)

            elif (random_sequence_type.upper() == 'FLOAT'):
                for i in range(length_of_value):
                    key += str(float(random.choice(string.digits)))

            elif (random_sequence_type.upper() == 'STRING'):
                for i in range(length_of_value):
                    key += random.choice(string.ascii_lowercase)

            elif (random_sequence_type.upper() == 'BOOLEAN'):
                words = ['TRUE', 'FALSE']
                key = random.choice(words)

            elif (random_sequence_type.upper() == 'OTHER'):
                if sequence_required == None:
                    words = []
                else:
                    words = sequence_required
                    key = random.choice(words)
            elif (random_sequence_type.upper() == 'EMAIL'):
                words = sequence_required
                key = random.choice(words) + "+" + random.choice(
                    self.random_sequences("NUMERIC", length_of_value, 5)) + random.choice(['@gmail.com', '@yahoo.com'])
            elif (random_sequence_type.upper() == 'GENERALNAME'):  # to create fake names(NAme21), address lines(AL12)

                key = string_to_be_concat + "_" + random.choice(self.random_sequences("NUMERIC", length_of_value, 100))
            elif (random_sequence_type.upper() == 'DATE'):
                new_time = datetime.datetime.now() - datetime.timedelta(
                    days=length_of_value * 365)  # Please provide the value in Years
                current_time = datetime.datetime.now()
                key = self.random_date(str(new_time.year) + "/1/1 12:00:00 AM",
                                       str(current_time.year) + "/1/1 12:00:00 AM", random.random())

            else:
                logger.info("Please provide the right sequence")
            counter = counter + 1
            values_of_the_list.append(key)
        return values_of_the_list

    def cmt_compare_date(self, dict_to_from_sfmc, dict_of_the_records_from_csv):
        sfmc_date = self.date_format_convert(dict_to_from_sfmc, 3)
        csv_date = self.date_format_convert(dict_of_the_records_from_csv, 4)
        if sfmc_date == csv_date:
            return True
        else:
            return False

    def form_dict_of_label_names_and_values_from_csv(self, work_sheet_object, record_identifier, column_header):
        common_utils = CommonUtilsClass()
        dict_of_the_records_from_csv = common_utils.get_row_values_with_column_header(work_sheet_object,
                                                                                      record_identifier, column_header)
        return dict_of_the_records_from_csv


    def get_all_row_values_for_column_header(self, work_sheet_object, cell_value):
        column_value = self.get_column_number_with_cell_value(work_sheet_object, cell_value)
        max_no_of_rows = work_sheet_object.max_row
        column_list = []
        for (row_counter) in range(2, max_no_of_rows + 1):  # to avoid the header
            column_list.append(unicode(work_sheet_object.cell(row=row_counter, column=column_value).value))
        return column_list

    #
    def get_row_values_with_column_header(self, work_sheet_object, row_value, column_header):
        '''
        this function creates a dictionary with key as column header and value will be cell
        value of the corresponding cell
        :param work_sheet_object:
        :param row_value:
        :param column_header:
        :return:
        '''
        row_number = self.get_row_number_with_cellvalue_col_header(work_sheet_object, row_value, column_header)
        max_no_of_columns = work_sheet_object.max_column
        row_dict = {}
        for (column_count) in range(1, max_no_of_columns + 1):
            column = unicode(work_sheet_object.cell(row=1, column=column_count).value)
            cell_value = unicode(work_sheet_object.cell(row=row_number, column=column_count).value)
            if cell_value == 'None':
                cell_value = ""
            row_dict[column.upper()] = unicode(cell_value.upper())
        return row_dict

    # to get the column headers list in a csv files
    def all_columns_names(self, file_path):
        with open(file_path, "r") as f:
            readers = unicodecsv.reader(f)
            return next(readers)

    # Common
    # Time stamp format "yyyymmdd"
    def last_user_created(self):
        '''
        Returns the TImeStamp of the current system time
        '''
        time = datetime.datetime.now()
        new_time = str.replace(str(time), '-', '')
        new_time = str.replace(str(new_time), ' ', '')
        new_time = str.replace(str(new_time), ':', '')
        return new_time[0:8]

    # Time stamp format "yyyymmdd"
    def test_time_stamp(self):
        '''
        Returns the TImeStamp of the current system time
        '''
        time = datetime.datetime.now()
        new_time = str.replace(str(time), '-', '')
        new_time = str.replace(str(new_time), ' ', '')
        new_time = str.replace(str(new_time), ':', '')
        #This is for few markets
        return new_time[0:8]

    # to get row value for a given cell value
    def get_row_number_with_cell_value(self, work_sheet_object, cell_value):
        max_no_of_columns = work_sheet_object.max_column
        max_no_of_rows = work_sheet_object.max_row
        for (row_counter) in range(1, max_no_of_rows + 1):
            for (column_counter) in range(1, max_no_of_columns + 1):
                if unicode(work_sheet_object.cell(row=row_counter, column=column_counter).value) == unicode(cell_value):
                    return row_counter

    # to get column value for a given cell value
    def get_column_number_with_cell_value(self, work_sheet_object, cell_value):
        max_no_of_columns = work_sheet_object.max_column
        max_no_of_rows = work_sheet_object.max_row
        for (row_counter) in range(1, max_no_of_rows + 1):
            for (column_counter) in range(1, max_no_of_columns + 1):
                if str(work_sheet_object.cell(row=row_counter, column=column_counter).value).upper() == str(cell_value.upper()):
                    return column_counter

    def create_csv_file_with_column_headers(self, work_sheet, table_headers_list):
        # Create the Column header for the  CSV file
        for column_counter in range(0, len(table_headers_list)):
            al = Alignment(horizontal="center", vertical="center")
            work_sheet.cell(row=1, column=column_counter + 1).alignment = al
            work_sheet.cell(row=1, column=column_counter + 1).value = table_headers_list[column_counter]

    def set_column_value_with_a_row(self, work_sheet, no_of_records, row_name, value_list):
        for row_counter in range(0, work_sheet.max_column + 1):
            if work_sheet.cell(row=1, column=row_counter + 1).value == row_name:
                for col_counter in range(0, no_of_records):
                    al = Alignment(horizontal="center", vertical="center")
                    work_sheet.cell(row=col_counter + 2, column=row_counter + 1).alignment = al
                    work_sheet.cell(row=col_counter + 2, column=row_counter + 1).value = value_list[col_counter]

    def get_value_based_on_col_header_row_header(self, work_sheet_object, row_header_value, column_header):
        row_value = self.get_row_number_with_cell_value(work_sheet_object, row_header_value)
        column_value = self.get_column_number_with_cell_value(work_sheet_object, column_header)
        return work_sheet_object.cell(row=row_value, column=column_value).value

    def get_row_number_with_cellvalue_col_header(self,work_sheet_object, cell_value, column_header):
        column_value = self.get_column_number_with_cell_value(work_sheet_object, column_header)
        max_no_of_rows = work_sheet_object.max_row
        for (row_counter) in range(1, max_no_of_rows + 1):
            if str(work_sheet_object.cell(row=row_counter, column=column_value).value) == str(cell_value):
                return row_counter

    def replace_cell_value_in_sheet(self, work_sheet_object, cell_existing_value, value_to_replace):
        max_no_of_columns = work_sheet_object.max_column
        max_no_of_rows = work_sheet_object.max_row
        for (row_counter) in range(1, max_no_of_rows + 1):
            for (column_counter) in range(1, max_no_of_columns + 1):
                if str(work_sheet_object.cell(row=row_counter, column=column_counter).value) == str(
                        cell_existing_value):
                    work_sheet_object.cell(row=row_counter, column=column_counter).value = value_to_replace

    # provide the ExcelName without xlsx
    # this function will create the fiven csv file sheets to new excel file

    def convert_xl_to_csv(self, excel_file, destination_filepath):
        '''
        Comments :: This function will create the each sheet of given excel to csv files
        :param excel_file:
        :param destination_filepath:
        :return:
        '''
        workbook = load_workbook(excel_file, data_only=True)
        sheets = workbook.sheetnames
        for worksheet_name in sheets:
            try:
                worksheet = workbook[worksheet_name]
            except KeyError:
                logger.info("Could not find " + worksheet_name)
                sys.exit(1)
            with open(destination_filepath + '\\' + worksheet_name + '.csv', 'wb') as f:
                c = unicodecsv.writer(f)
                for r in worksheet.rows:
                    cell_write_list = []
                    for cell in r:
                        if unicode(cell.value).encode("utf-8") == 'None':
                            cell_write_list.append("")
                        else:
                            cell_write_list.append(unicode(cell.value).encode("utf-8"))
                    c.writerow(cell_write_list)
            logger.info("Created " + worksheet_name + "worksheet in " + destination_filepath + "file path")


    def convert_csv_to_new_excel(self, source_filepath, csv_filename, excel_file_name, destination_path=None):
        work_book_object = Workbook()
        if ".xlsx" in excel_file_name:
            excel_file_name.replace(".xlsx", "")
        work_sheet_object = work_book_object.create_sheet(title=excel_file_name)
        csv_filepath = source_filepath + csv_filename
        csv_file = open(csv_filepath)
        reader = unicodecsv.reader(csv_file, encoding='utf-8', errors='ignore', delimiter=',')
        for row_counter in reader:
            [i.encode('utf-8') for i in row_counter]
            work_sheet_object.append(row_counter)
        if destination_path == None:
            destination_path = source_filepath
        excel_file_path = destination_path + excel_file_name + ".xlsx"
        work_book_object.save(excel_file_path)

    def get_csv_file_names_from_folder(self, cmt_table_template_csv):
        '''
        to get the CSV file names in particular folder name
        :param cmt_table_template_csv:
        :return:
        '''
        commonutils = CommonUtilsClass()
        csv_file_list = [f for f in listdir(cmt_table_template_csv) if isfile(join(cmt_table_template_csv, f))]
        all_tables_headers_dict = {}
        for file_object in csv_file_list:
            all_columns = commonutils.all_columns_names(cmt_table_template_csv + "\\" + file_object)
            all_tables_headers_dict[(file_object.upper()).replace(".CSV", "")] = all_columns
        return all_tables_headers_dict

    def convert_n_add_csv_to_existing_excel(self, source_filepath_of_csv, csv_filename, excel_file_path):
        work_book_object = load_workbook(excel_file_path)
        csv_filename_with_timestamp = csv_filename
        try:
            work_sheet_object = work_book_object.get_sheet_by_name(csv_filename_with_timestamp)
            logger.info(csv_filename_with_timestamp + " sheet already exists in the Excel")
            work_sheet_object.remove_sheet(csv_filename_with_timestamp)
        except:
            logger.info(csv_filename_with_timestamp + " sheet doesn't exist in the Excel")
        work_sheet_object = work_book_object.create_sheet(title=csv_filename_with_timestamp)
        csv_filepath = source_filepath_of_csv + csv_filename
        if not ".csv" in csv_filepath:
            csv_filepath = csv_filepath + ".csv"
        csv_file = open(csv_filepath)
        reader = unicodecsv.reader(csv_file, encoding='utf-8', errors='ignore', delimiter=',')
        for row_counter in reader:
            [i.encode('utf-8') for i in row_counter]
            work_sheet_object.append(row_counter)
        work_book_object.save(excel_file_path)

    def get_work_sheet_object(self, file_path, sheet_name):
        workbook = load_workbook(file_path)
        worksheetname = workbook[sheet_name]
        return worksheetname

    def create_predefined_attributes(self, work_sheet, no_of_records, table_headers_list, attribute_name, attributes,
                                     random_type=None):
        if random_type == "OFF":
            random_sequen_list = attributes
        else:
            random_sequen_list = self.random_sequences('OTHER', 32, no_of_records, attributes)
        for column_counter in range(0, len(table_headers_list)):
            if work_sheet.cell(row=1, column=column_counter + 1).value == attribute_name:
                for row_counter in range(0, no_of_records):
                    al = Alignment(horizontal="center", vertical="center")
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).alignment = al
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).value = random_sequen_list[
                        row_counter]

    def get_attribute_values(self, work_sheet, table_headers_list, attribute_name):
        key = []
        for column_counter in range(0, len(table_headers_list)):
            if work_sheet.cell(row=1, column=column_counter + 1).value == attribute_name:
                for row_counter in range(2, work_sheet.max_row + 1):
                    key.append(work_sheet.cell(row=row_counter, column=column_counter + 1).value)
                return key

    def create_email_attributes(self, work_sheet, no_of_records, table_headers_list, attribute_name, string_length,
                                emailaddress):
        random_sequen_list = self.random_sequences('EMAIL', string_length, no_of_records, emailaddress)
        for column_counter in range(0, len(table_headers_list)):
            if work_sheet.cell(row=1, column=column_counter + 1).value == attribute_name:
                for row_counter in range(0, no_of_records):
                    al = Alignment(horizontal="center", vertical="center")
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).alignment = al
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).value = random_sequen_list[
                        row_counter]

    def create_alpha_numeric_attributes(self, work_sheet, no_of_records, table_headers_list, attribute_name,
                                        string_length):
        random_sequen_list = self.random_sequences('ALPHANUMERIC', string_length, no_of_records)
        for column_counter in range(0, len(table_headers_list)):
            if work_sheet.cell(row=1, column=column_counter + 1).value == attribute_name:
                for row_counter in range(0, no_of_records):
                    al = Alignment(horizontal="center", vertical="center")
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).alignment = al
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).value = random_sequen_list[
                        row_counter]

    def create_string_attributes(self, work_sheet, no_of_records, table_headers_list, attribute_name, string_length,
                                 random_type="STRING", string_to_be_concat=None):
        random_sequen_list = self.random_sequences(random_type, string_length, no_of_records,
                                                   string_to_be_concat=string_to_be_concat)
        for column_counter in range(0, len(table_headers_list)):
            if work_sheet.cell(row=1, column=column_counter + 1).value == attribute_name:
                for row_counter in range(0, no_of_records):
                    al = Alignment(horizontal="center", vertical="center")
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).alignment = al
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).value = random_sequen_list[
                        row_counter]

    def create_boolean_attributes(self, work_sheet, no_of_records, table_headers_list, attribute_name):
        random_sequen_list = self.random_sequences('BOOLEAN', "", no_of_records)
        for column_counter in range(0, len(table_headers_list)):
            if work_sheet.cell(row=1, column=column_counter + 1).value == attribute_name:
                for row_counter in range(0, no_of_records):
                    al = Alignment(horizontal="center", vertical="center")
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).alignment = al
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).value = random_sequen_list[
                        row_counter]

    def create_nuermic_attributes(self, work_sheet, no_of_records, table_headers_list, header_name, length_of_value):
        random_sequen_list = self.random_sequences('NUMERIC', length_of_value, no_of_records)
        for column_counter in range(0, len(table_headers_list)):
            if work_sheet.cell(row=1, column=column_counter + 1).value == header_name:
                for row_counter in range(0, no_of_records):
                    al = Alignment(horizontal="center", vertical="center")
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).alignment = al
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).value = random_sequen_list[
                        row_counter]

    def create_float_attributes(self, work_sheet, no_of_records, table_headers_list, header_name, length_of_value):
        random_sequen_list = self.random_sequences('FLOAT', length_of_value, no_of_records)
        for column_counter in range(0, len(table_headers_list)):
            if work_sheet.cell(row=1, column=column_counter + 1).value == header_name:
                for row_counter in range(0, no_of_records):
                    al = Alignment(horizontal="center", vertical="center")
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).alignment = al
                    work_sheet.cell(row=row_counter + 2, column=column_counter + 1).value = random_sequen_list[
                        row_counter]

    def make_first_char_upper_case(self, strng_in_question):
        all_lower = strng_in_question.lower()
        fc_upper = all_lower[0].upper() + all_lower[1:]
        return fc_upper

    # Example Input ="aa bb cc dd"  output="Aa Bb Cc Dd"
    def make_first_char_upper_for_all_words_in_string(self, strng_in_question):
        strng_in_question_list = strng_in_question.split(" ")
        string_with_all_first_word_char_upper_case = ""
        for word in strng_in_question_list:
            string_with_all_first_word_char_upper_case = string_with_all_first_word_char_upper_case + " " + \
                                                         self.first_char_higher_case_string(word)
        return string_with_all_first_word_char_upper_case[1:len(string_with_all_first_word_char_upper_case)]

    def update_sheetname_timestamp(self, excel_path):
        wrkbook = load_workbook(excel_path)
        sheet_list = wrkbook.sheetnames
        for sheetname in sheet_list:
            wrksheet = wrkbook[sheetname]
            wrksheet.title = sheetname[0:-8]
            wrkbook.save(excel_path)

    def all_cases_words_in_sentences(self, string_to_extract):
        '''
        Eg Input :: staging in put
            output :: ['Staging in put', 'Staging In put', 'Staging In Put', 'Staging in Put', 'staging in Put',
            'staging in put', 'staging In put', 'staging In Put']
        :param string_to_extract:
        :return:
        '''
        all_cases_words_in_sentences = [string_to_extract]
        string_combinations = {}
        count = 0
        for string in string_to_extract.split(" "):
            if self.check_if_string_has_only_alphabets(string):
                item__select_format1 = self.make_string_all_lower(string)  # all lower  eg "staging opt in"
                item__select_format2 = self.make_first_char_upper_case(string)  # all lower  eg "Staging"
                string_combinations[string] = [item__select_format1, item__select_format2]
                count = count + 1
            else:
                string_combinations[string] = [string]
        string = string_to_extract.split(" ")
        while (len(all_cases_words_in_sentences) < pow(2, count)):
            temp = ""
            for x in range(0, len(string)):
                temp = temp + " " + random.choice(string_combinations[string[x]])                
            temp = temp[1:len(temp)]
            if temp not in all_cases_words_in_sentences:
                all_cases_words_in_sentences.append(temp)
        return all_cases_words_in_sentences

    def make_string_all_upper(self, strng_in_question):
        all_upp = strng_in_question.upper()
        return all_upp

    def remove_spaces_in_string(self, strng_in_question):
        strng_in_question_list = strng_in_question.split(" ")
        string_without_spaces = ""
        for word in strng_in_question_list:
            string_without_spaces = string_without_spaces + word
        return string_without_spaces

    def make_string_all_lower(self, strng_in_question):
        all_low = strng_in_question.lower()
        return all_low

    def delete_a_file(self, filepath):
        os.remove(filepath)

    def delete_selected_files_from_a_directory(self, dir_path, file_xtension):
        files_path = dir_path + "/*." + file_xtension
        files = glob.glob(files_path)
        for f in files:
            logger.info("Removed file ::" + f)
            os.remove(f)

    def delete_only_folder_of_a_directory(self, dir_path):
        for root, dirs, files in os.walk(dir_path):
            for d in dirs:
                shutil.rmtree(os.path.join(root, d))

    def first_char_lower_case_string(self, string_in_question):
        first_char = string_in_question[:1]
        onlyaplha = self.check_if_string_has_only_alphabets(first_char)
        if (onlyaplha):
            f_char_lower = lambda s: s[:1].lower() + s[1:] if s else ''
            first_char_lower = f_char_lower(string_in_question)
            return first_char_lower
        else:
            logger.info(first_char + "  is not alphabet")

    def first_char_higher_case_string(self, string_in_question):
        first_char = string_in_question[:1]
        onlyaplha = self.check_if_string_has_only_alphabets(first_char)
        if (onlyaplha):
            f_char_higher = lambda s: s[:1].upper() + s[1:].lower() if s else ''
            first_char_higher = f_char_higher(string_in_question)
            return first_char_higher
        else:
            logger.info(first_char + "  is not alphabet")

    def check_if_string_has_only_alphabets(self, string_to_be_checked):
        check_re_pat = re.compile('[a-zA-Z]')
        return check_re_pat.match(string_to_be_checked)

    # remove a value from json
    def remove_key_json_code(self, json_code, key_to_remove_list):
        json_code = json.loads(json_code,
                               encoding='utf-8')  # Input jason code will be string format converting it into dictionary
        keys_json_code = json_code['KEYS']
        values_json_code = json_code['VALUES']
        for key_to_remove in key_to_remove_list:
            if key_to_remove.upper() in keys_json_code:
                del (keys_json_code[key_to_remove.upper()])
            else:
                del (values_json_code[key_to_remove])
        return json.dumps(json_code)

    # replace the attribute all attribute values
    def replace_attribute_value_in_jason_code(self, json_code, attribute_list, value_to_assign):
        json_code = json.loads(json_code,
                               encoding='utf-8')  # Input jason code will be string format converting it into dictionary
        keys_json_code = json_code['KEYS']
        values_json_code = json_code['VALUES']
        for attribute in attribute_list:
            for key in keys_json_code:
                if key.upper() == attribute.upper():
                    keys_json_code[attribute.upper()] = unicode(value_to_assign)
                    break
            for value in values_json_code:
                if value.upper() == attribute.upper():
                    values_json_code[value] = unicode(value_to_assign)
                    break
        return json.dumps(json_code)  # dictionary to string format

    def copy_contents_from_file_to_file(self, source_filepath, destination_filepath):
        data_to_insert = self.read_data_from_file(source_filepath)
        self.write_data_into_file(destination_filepath, data_to_insert)

    def write_data_into_file(self, filepath, data_to_insert):
        file_object = open(filepath, 'w+')
        file_object.flush()
        file_object.write(data_to_insert)
        file_object.close()

    def read_data_from_file(self, filepath):
        file_object = open(filepath, 'r+')
        data_to_read = file_object.read()
        file_object.close()
        return data_to_read

    def copy_file(self, source_locations, destination_location):
        '''
        Please specify the source_locations, destination_location with path and file name attached
        :param source_locations:
        :param destination_location:
        :return:
        '''
        shutil.copy(source_locations, destination_location)

    # ======================= I/O Parameters Functions (Start) ===========================
    def set_Value_In_IOparmetres(self, parameter_key, parameter_value):
        section_name = "IOParameterSection"
        config_prop_f_name = "IOParameters.properties"
        self.set_val_for_key_in_config_propfile(config_prop_f_name, section_name,
                                                parameter_key, parameter_value)
        logger.info("Successfully set the '" + str(parameter_key) + "' key with '" + str(
            parameter_value) + "' value in to the I/O parameter file")

    def get_Value_In_IOparmetres(self, parameter_key):
        section_name = "IOParameterSection"
        config_prop_f_name = "IOParameters.properties"
        return self.get_val_from_config_propfile(config_prop_f_name, section_name,
                                                                  parameter_key)

    # ======================= I/O Parameters Functions (End) ===========================

    # ========================= Base64 Encoding and Decoding ===========================
    def base64_encode(self, encryption_key, input_string):
        enc = []
        for i in range(len(input_string)):
            key_c = encryption_key[i % len(encryption_key)]
            enc_c = chr((ord(input_string[i]) + ord(key_c)) % 256)
            enc.append(enc_c)
        return base64.urlsafe_b64encode("".join(enc))

    def base64_decode(self, encryption_key, input_string):
        dec = []
        input_string = base64.urlsafe_b64decode(input_string.encode('ascii', 'ignore'))
        for i in range(len(input_string)):
            key_c = encryption_key[i % len(encryption_key)]
            dec_c = chr((256 + ord(input_string[i]) - ord(key_c)) % 256)
            dec.append(dec_c)
        return "".join(dec)
    # ========================= Base64 Encoding and Decoding ============================

    # ========================= AES256 Encryption and Decryption ============================
    def AES256_encrypt(self, input_string, key):
        private_key = hashlib.sha256(key.encode("utf-8")).digest()
        raw = pad(input_string)
        iv = Random.new().read(AES.block_size)
        cipher = AES.new(private_key, AES.MODE_CBC, iv)
        return base64.b64encode(iv + cipher.encrypt(raw))

    def AES256_decrypt(self, key, encrypted_string):
        private_key = hashlib.sha256(key.encode("utf-8")).digest()
        enc = base64.b64decode(encrypted_string)
        iv = enc[:16]
        cipher = AES.new(private_key, AES.MODE_CBC, iv)
        return unpad(cipher.decrypt(enc[16:]))
    # ========================= AES256 Encryption and Decryption ============================

    # ========================= Chrome Options ==========================================
    def build_chrome_options(self):
        chrome_options = webdriver.ChromeOptions()
        chrome_options.accept_untrusted_certs = True
        chrome_options.assume_untrusted_cert_issuer = True
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-impl-side-painting")
        chrome_options.add_argument("--disable-setuid-sandbox")
        chrome_options.add_argument("--disable-seccomp-filter-sandbox")
        chrome_options.add_argument("--disable-breakpad")
        chrome_options.add_argument("--disable-client-side-phishing-detection")
        chrome_options.add_argument("--disable-cast")
        chrome_options.add_argument("--disable-cast-streaming-hw-encoding")
        chrome_options.add_argument("--disable-cloud-import")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--disable-session-crashed-bubble")
        chrome_options.add_argument("--disable-ipv6")
        chrome_options.add_argument("--allow-http-screen-capture")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_experimental_option('useAutomationExtension', False)
        return chrome_options

# ==============================================  (End of the File) ==================================================